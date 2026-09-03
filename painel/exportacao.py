from io import BytesIO

from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from painel.utils import filtrar_historico_por_periodo

COLUNAS = ["Data do desligamento", "Setor", "Colaborador", "Cargo"]
COLUNAS_ADMISSOES = ["Data da admissão", "Setor", "Colaborador", "Cargo"]


def _cabecalho_padrao(aba, colunas):
    aba.append(colunas)
    fonte_cabecalho = Font(name="Arial", bold=True, color="FFFFFF")
    preenchimento_cabecalho = PatternFill(start_color="00995D", end_color="00995D", fill_type="solid")
    for celula in aba[1]:
        celula.font = fonte_cabecalho
        celula.fill = preenchimento_cabecalho
        celula.alignment = Alignment(vertical="center")


def gerar_planilha_admissoes(rows, inicio=None, fim=None):
    linhas = []
    for linha in rows:
        ts = linha.get("admission_ts")
        if not ts:
            continue
        data_iso = datetime.fromtimestamp(ts / 1000, tz=timezone.utc).strftime("%Y-%m-%d")
        if inicio and data_iso < inicio:
            continue
        if fim and data_iso > fim:
            continue
        linhas.append(linha)

    linhas.sort(key=lambda linha: (
        (linha.get("workplaces") or "").split(",")[0].strip() or "Não informado",
        linha.get("admission_ts") or 0,
    ))

    workbook = Workbook()
    aba = workbook.active
    aba.title = "Admissões por setor"
    _cabecalho_padrao(aba, COLUNAS_ADMISSOES)

    for linha in linhas:
        aba.append([
            linha.get("admission_date_str") or "",
            (linha.get("workplaces") or "").split(",")[0].strip() or "Não informado",
            linha.get("name") or "",
            linha.get("job_role") or "",
        ])

    for coluna in ["A", "B", "C", "D"]:
        for celula in aba[coluna][1:]:
            celula.font = Font(name="Arial")

    larguras = [20, 26, 32, 26]
    for indice, largura in enumerate(larguras, start=1):
        aba.column_dimensions[get_column_letter(indice)].width = largura

    aba.freeze_panes = "A2"
    aba.auto_filter.ref = f"A1:D{aba.max_row}"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def gerar_planilha_desligamentos(historico, inicio=None, fim=None):
    linhas = filtrar_historico_por_periodo(historico, inicio, fim)
    linhas.sort(key=lambda linha: (
        (linha.get("setor") or "").strip() or "Não informado",
        linha.get("data_registro") or "",
    ))

    workbook = Workbook()
    aba = workbook.active
    aba.title = "Desligamentos por setor"
    _cabecalho_padrao(aba, COLUNAS)

    for linha in linhas:
        aba.append([
            linha.get("data_desligamento") or "",
            (linha.get("setor") or "").strip() or "Não informado",
            linha.get("nome_colaborador") or "",
            linha.get("cargo") or "",
        ])

    for celula in aba["A"][1:]:
        celula.font = Font(name="Arial")
    for coluna in ["B", "C", "D"]:
        for celula in aba[coluna][1:]:
            celula.font = Font(name="Arial")

    larguras = [20, 26, 32, 26]
    for indice, largura in enumerate(larguras, start=1):
        aba.column_dimensions[get_column_letter(indice)].width = largura

    aba.freeze_panes = "A2"
    aba.auto_filter.ref = f"A1:D{aba.max_row}"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
